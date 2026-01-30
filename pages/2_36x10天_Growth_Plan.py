from datetime import date
import io
import streamlit as st

# ✅ 必须在任何 st.xxx 之前
lang = st.session_state.get("lang", "zh")
st.set_page_config(
    page_title=("② 36×10：自我提升计划" if lang == "zh" else "② 36×10: Growth Plan"),
    page_icon="🌱",
    layout="wide",
)

from i18n import init_i18n, lang_selector, t

from db import (
    regenerate_sprints,
    get_sprints,
    get_sprint_by_no,
    update_sprint_text,
    list_tasks_for_sprint,
    add_task_to_sprint_unique,
    toggle_task_done,
    update_task_evidence,
)

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# -----------------------
# 兼容旧版 Streamlit：统一 rerun
# -----------------------
def rerun():
    if hasattr(st, "rerun"):
        st.rerun()
    else:
        st.experimental_rerun()


# -----------------------
# i18n 初始化 + 侧边栏语言
# -----------------------
init_i18n(default="zh")
lang_selector()


# -----------------------
# 36×10 Excel 导出：6×6 大表
# -----------------------
def build_36x10_excel() -> bytes:
    """
    生成一个好看的 6×6 大表：
    - 每个周期块：表头=主题；交付物；任务列表（含完成状态）
    """
    periods = get_sprints()
    if not periods:
        return b""

    BLOCK_COLS = 3
    BLOCK_ROWS = 10
    GAP_COL = 1
    GAP_ROW = 1

    wb = Workbook()
    ws = wb.active
    ws.title = "36×10 Plan" if st.session_state.lang == "en" else "36×10 自我提升计划"

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    font_header = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    font_body = Font(name="Microsoft YaHei", size=10, color="1F1F1F")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    fill_header = PatternFill("solid", fgColor="6C5CE7")
    fill_obj = PatternFill("solid", fgColor="F7F7FB")
    fill_task = PatternFill("solid", fgColor="FFFFFF")
    fill_done = PatternFill("solid", fgColor="E9F7EF")

    # Title row
    total_cols = 6 * (BLOCK_COLS + GAP_COL)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_text = "36×10 Growth Plan (6×6 Master Sheet)" if st.session_state.lang == "en" else "36×10 自我提升计划（6×6 大表）"
    tcell = ws.cell(row=1, column=1, value=title_text)
    tcell.font = Font(name="Microsoft YaHei", bold=True, size=16, color="1F1F1F")
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Column widths
    for c in range(1, total_cols + 1):
        letter = get_column_letter(c)
        if (c % (BLOCK_COLS + GAP_COL)) == 0:
            ws.column_dimensions[letter].width = 3
        else:
            ws.column_dimensions[letter].width = 18

    for r in range(2, 2 + 6 * (BLOCK_ROWS + GAP_ROW) + 2):
        ws.row_dimensions[r].height = 18

    periods_sorted = sorted(periods, key=lambda x: x.sprint_no)

    def top_left_of_block(sprint_no: int):
        idx = sprint_no - 1
        block_r = idx // 6
        block_c = idx % 6
        start_row = 2 + block_r * (BLOCK_ROWS + GAP_ROW)
        start_col = 1 + block_c * (BLOCK_COLS + GAP_COL)
        return start_row, start_col

    def merge_block(row, col, r_span, c_span):
        ws.merge_cells(start_row=row, start_column=col, end_row=row + r_span - 1, end_column=col + c_span - 1)

    def set_block_border(r0, c0, r1, c1):
        for rr in range(r0, r1 + 1):
            for cc in range(c0, c1 + 1):
                ws.cell(rr, cc).border = border

    deliverable_label = "Deliverables:\n" if st.session_state.lang == "en" else "交付物/成果：\n"

    for sp in periods_sorted:
        r0, c0 = top_left_of_block(sp.sprint_no)
        r1 = r0 + BLOCK_ROWS - 1
        c1 = c0 + BLOCK_COLS - 1

        theme = (sp.theme or "").strip()
        header_text = theme if theme else ("Untitled" if st.session_state.lang == "en" else "未命名主题")
        header_text = (f"Cycle {sp.sprint_no} | {header_text}" if st.session_state.lang == "en"
                       else f"第{sp.sprint_no}周期｜{header_text}")

        merge_block(r0, c0, 1, BLOCK_COLS)
        hc = ws.cell(r0, c0, header_text)
        hc.font = font_header
        hc.fill = fill_header
        hc.alignment = align_center
        ws.row_dimensions[r0].height = 26

        obj = (sp.objective or "").strip()
        obj_text = obj if obj else ("(Not set)" if st.session_state.lang == "en" else "（未填写交付物）")
        merge_block(r0 + 1, c0, 2, BLOCK_COLS)
        oc = ws.cell(r0 + 1, c0, f"{deliverable_label}{obj_text}")
        oc.font = font_body
        oc.fill = fill_obj
        oc.alignment = align_left
        ws.row_dimensions[r0 + 1].height = 38
        ws.row_dimensions[r0 + 2].height = 38

        tasks = list_tasks_for_sprint(sp.sprint_no)
        max_tasks = 6
        show = tasks[:max_tasks]
        more = max(0, len(tasks) - len(show))

        for i in range(max_tasks):
            rr = r0 + 3 + i
            merge_block(rr, c0, 1, BLOCK_COLS)
            if i < len(show):
                tt = show[i]
                mark = "✅" if tt.done else "⬜"
                txt = f"{mark} {tt.title}"
                cell = ws.cell(rr, c0, txt)
                cell.fill = fill_done if tt.done else fill_task
            else:
                cell = ws.cell(rr, c0, "")
                cell.fill = fill_task

            cell.font = font_body
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[rr].height = 20

        rr_hint = r0 + 9
        merge_block(rr_hint, c0, 1, BLOCK_COLS)
        hint = (f"… {more} more tasks" if st.session_state.lang == "en" else f"…还有 {more} 条任务") if more > 0 else ""
        hint_cell = ws.cell(rr_hint, c0, hint)
        hint_cell.font = Font(name="Microsoft YaHei", size=9, color="666666", italic=True)
        hint_cell.alignment = Alignment(horizontal="right", vertical="center")
        hint_cell.fill = fill_task
        ws.row_dimensions[rr_hint].height = 18

        set_block_border(r0, c0, r1, c1)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# -----------------------
# 页面样式
# -----------------------
st.markdown(
    """
<style>
.block-container { padding-top: 1.4rem; padding-bottom: 2.0rem; max-width: 1180px; }
.card {
    background: #fff;
    border-radius: 16px;
    padding: 14px 14px;
    margin-bottom: 12px;
    border: 1px solid rgba(0,0,0,0.06);
    box-shadow: 0 10px 24px rgba(0,0,0,0.04);
}
.small { color:#666; font-size: 13px; }
</style>
""",
    unsafe_allow_html=True,
)

# -----------------------
# 页面标题（双语）
# -----------------------
st.title(t("page_36_title"))
st.caption(t("page_36_caption"))

# Session
if "current_sprint_no" not in st.session_state:
    st.session_state.current_sprint_no = None

# -----------------------
# 生成周期（双语）
# -----------------------
with st.expander(
    "首次使用：生成 36 个「10天行动周期」（建议只做一次）"
    if st.session_state.lang == "zh"
    else "First time: generate 36 ten-day cycles (recommended once)",
    expanded=False,
):
    start = st.date_input(
        "请选择开始日期" if st.session_state.lang == "zh" else "Choose start date",
        value=date.today(),
    )
    if st.button(
        "🚀 生成/重建 36×10（会清空旧周期与任务）"
        if st.session_state.lang == "zh"
        else "🚀 Generate/Rebuild 36×10 (will clear old cycles & tasks)",
        use_container_width=True,
    ):
        regenerate_sprints(start)
        st.success("已生成 36 个「10天行动周期」✅" if st.session_state.lang == "zh" else "Generated 36 cycles ✅")
        st.session_state.current_sprint_no = 1
        rerun()

periods = get_sprints()
if not periods:
    st.info("还没有行动周期。请先在上面生成 36×10。" if st.session_state.lang == "zh"
            else "No cycles yet. Please generate 36×10 above.")
    st.stop()

# -----------------------
# 导出按钮（双语）
# -----------------------
st.markdown("### " + ("导出" if st.session_state.lang == "zh" else "Export"))
xlsx_bytes = build_36x10_excel()
filename = "36x10_plan.xlsx" if st.session_state.lang == "en" else "36x10_自我提升计划.xlsx"
st.download_button(
    t("export_excel_btn"),
    data=xlsx_bytes,
    file_name=filename,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

st.divider()


# -----------------------
# 详情视图
# -----------------------
def render_period_detail(period_no: int):
    s = get_sprint_by_no(period_no)
    if not s:
        st.error("找不到该行动周期" if st.session_state.lang == "zh" else "Cycle not found")
        return

    title = (f"第 {s.sprint_no} 个 10天行动周期（详情）"
             if st.session_state.lang == "zh"
             else f"Cycle {s.sprint_no} (Details)")
    st.markdown(f"## {title}")
    st.write(
        (f"日期：{s.start_date} ~ {s.end_date}" if st.session_state.lang == "zh"
         else f"Dates: {s.start_date} ~ {s.end_date}")
    )

    with st.form(f"period_edit_{period_no}"):
        theme = st.text_input(
            "主题（这10天主要围绕什么？）" if st.session_state.lang == "zh" else "Theme (what is this cycle about?)",
            value=s.theme,
            placeholder=("例如：规则制定与对外表达" if st.session_state.lang == "zh" else "e.g., Messaging & external communication"),
        )

        objective = st.text_area(
            "本周期成果/交付物（10天结束时，你要交付什么？越具体越好）"
            if st.session_state.lang == "zh"
            else "Deliverables (what will you deliver by the end of 10 days?)",
            value=s.objective,
            height=120,
            placeholder=(
                "例如：DT规则v1（文档）+ 对外发布说明文章（链接）+ 1个可展示Demo（截图/链接）"
                if st.session_state.lang == "zh"
                else "e.g., Rule doc v1 + public post (link) + 1 demo (screenshot/link)"
            ),
        )

        review = st.text_area(
            "复盘 Review（发生了什么？学到什么？下一步怎么改？）"
            if st.session_state.lang == "zh"
            else "Review (what happened? what did you learn? what to improve next?)",
            value=s.review,
            height=140,
            placeholder=("例如：最有效的是……下次把……提前" if st.session_state.lang == "zh" else "e.g., What worked… Next time, do X earlier…"),
        )

        ok = st.form_submit_button(
            "💾 保存（主题/交付物/复盘）" if st.session_state.lang == "zh" else "💾 Save (Theme/Deliverables/Review)",
            use_container_width=True,
        )

    if ok:
        update_sprint_text(period_no, theme, objective, review, mit="")  # mit 不用
        st.success("已保存 ✅" if st.session_state.lang == "zh" else "Saved ✅")
        rerun()

    st.markdown("---")
    st.subheader("任务列表（执行清单）" if st.session_state.lang == "zh" else "Task List")

    tasks = list_tasks_for_sprint(period_no)

    with st.form(f"add_task_{period_no}"):
        new_title = st.text_area(
            "新增任务（每行一个）" if st.session_state.lang == "zh" else "Add tasks (one per line)",
            value="",
            height=90,
            placeholder=("例如：完成DT规则v1\n例如：写并发布说明文章" if st.session_state.lang == "zh"
                        else "e.g., Finish rule doc v1\n e.g., Publish the post"),
        )
        add = st.form_submit_button("➕ 添加任务" if st.session_state.lang == "zh" else "➕ Add tasks", use_container_width=True)

    if add:
        if new_title.strip():
            for line in new_title.splitlines():
                line = line.strip()
                if line:
                    add_task_to_sprint_unique(period_no, line, source_care_id=None)
            st.success("已添加 ✅" if st.session_state.lang == "zh" else "Added ✅")
            rerun()

    if not tasks:
        st.info("该周期还没有任务。你可以从年度挖掘分配，或在这里新增。" if st.session_state.lang == "zh"
                else "No tasks yet. You can assign from Life Circle, or add them here.")
        return

    for tt in tasks:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        c1, c2 = st.columns([1, 3])

        with c1:
            checked = st.checkbox("完成" if st.session_state.lang == "zh" else "Done", value=bool(tt.done), key=f"done_{tt.id}")
            if checked != bool(tt.done):
                toggle_task_done(tt.id, checked)
                rerun()

        with c2:
            st.write(f"**{tt.title}**")
            ev = st.text_area(
                "证据/记录（链接、截图说明、复盘要点）" if st.session_state.lang == "zh" else "Evidence / Notes (links, screenshots, learnings)",
                value=tt.evidence or "",
                height=70,
                key=f"ev_{tt.id}",
            )
            if ev != (tt.evidence or ""):
                update_task_evidence(tt.id, ev)

        st.markdown("</div>", unsafe_allow_html=True)


# -----------------------
# 顶部导航（详情页）
# -----------------------
def goto_period(no: int):
    st.session_state.current_sprint_no = int(no)
    rerun()

def back_to_overview():
    st.session_state.current_sprint_no = None
    rerun()


if st.session_state.current_sprint_no is not None:
    top = st.columns([1, 2, 1])
    with top[0]:
        st.button("⬅ 返回总览" if st.session_state.lang == "zh" else "⬅ Back", use_container_width=True, on_click=back_to_overview)
    with top[2]:
        cur = int(st.session_state.current_sprint_no)
        st.button("← 上一个" if st.session_state.lang == "zh" else "← Prev", use_container_width=True, on_click=goto_period, args=(max(1, cur - 1),))
        st.button("下一个 →" if st.session_state.lang == "zh" else "Next →", use_container_width=True, on_click=goto_period, args=(min(36, cur + 1),))

    render_period_detail(int(st.session_state.current_sprint_no))
    st.stop()


# -----------------------
# 总览视图
# -----------------------
st.markdown("## " + ("36 个 10天行动周期 总览（点击进入）" if st.session_state.lang == "zh"
                    else "Overview: 36 Ten-Day Cycles (click to open)"))

st.caption("提示：年度挖掘分配后，第1~N个周期会自动出现任务。" if st.session_state.lang == "zh"
           else "Tip: after auto-assignment, cycles 1..N will contain tasks.")

def enter_period(no: int):
    st.session_state.current_sprint_no = int(no)
    rerun()

cols = st.columns(3)
periods_sorted = sorted(periods, key=lambda x: x.sprint_no)
for idx, sp in enumerate(periods_sorted):
    with cols[idx % 3]:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.write(f"**{'第' if st.session_state.lang == 'zh' else 'Cycle '} {sp.sprint_no} {'个 10天行动周期' if st.session_state.lang == 'zh' else ''}**")
        st.write(f"{sp.start_date} ~ {sp.end_date}")

        tasks = list_tasks_for_sprint(sp.sprint_no)
        done_cnt = sum(1 for x in tasks if x.done)
        st.caption(
            (f"任务完成：{done_cnt}/{len(tasks)}" if st.session_state.lang == "zh"
             else f"Done: {done_cnt}/{len(tasks)}")
        )

        if sp.theme:
            st.caption(("主题：" if st.session_state.lang == "zh" else "Theme: ") + sp.theme)

        if sp.objective:
            short_obj = sp.objective.strip().replace("\n", " ")
            if len(short_obj) > 40:
                short_obj = short_obj[:40] + "…"
            st.caption(("交付物：" if st.session_state.lang == "zh" else "Deliverables: ") + short_obj)

        st.button(
            "进入" if st.session_state.lang == "zh" else "Open",
            key=f"enter_{sp.sprint_no}",
            use_container_width=True,
            on_click=enter_period,
            args=(sp.sprint_no,),
        )
        st.markdown("</div>", unsafe_allow_html=True)
