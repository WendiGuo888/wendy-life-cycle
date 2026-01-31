# pages/4_导出_Export_Hub.py
# -*- coding: utf-8 -*-

import io
import json
import textwrap

import streamlit as st

# ✅ Matplotlib 在 Cloud 上建议用 Agg
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Circle

from i18n import init_i18n, lang_selector, t

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from store import (
    get_or_create_annual_dig,
    get_sprints,
    list_tasks_for_sprint,
    export_user_json,
    import_user_json,
)

# -----------------------
# ✅ set_page_config 必须在任何 st.xxx 前
# -----------------------
lang = st.session_state.get("lang", "zh")
st.set_page_config(
    page_title=("④ 导出中心" if lang == "zh" else "④ Export Hub"),
    page_icon="🌱",
    layout="wide",
)

# -----------------------
# i18n 初始化 + 侧边栏语言
# -----------------------
init_i18n(default="zh")
lang_selector()

# -----------------------
# 样式（✅ 顶部显示不全：padding-top 调大 + max-width 放宽）
# -----------------------
st.markdown(
    """
<style>
/* Cloud 上更稳：顶部多留一点空间，避免第一块内容“顶到天花板” */
.block-container { padding-top: 2.2rem; padding-bottom: 2.2rem; max-width: 1280px; }

/* 卡片样式 */
.card {
    background: #fff;
    border-radius: 16px;
    padding: 16px 16px;
    margin-bottom: 14px;
    border: 1px solid rgba(0,0,0,0.06);
    box-shadow: 0 10px 24px rgba(0,0,0,0.04);
}

.small { color:#666; font-size: 13px; }

/* expander 内部更像产品 */
[data-testid="stExpander"] { border-radius: 14px; }
</style>
""",
    unsafe_allow_html=True,
)

st.title(t("page_export_title"))
st.caption(t("page_export_caption"))

# =======================
# 工具函数
# =======================
def safe_load_json(s: str):
    try:
        return json.loads(s) if s else {}
    except Exception:
        return {}

def get_meta(intersections: dict) -> dict:
    meta = intersections.get("_meta", {})
    return meta if isinstance(meta, dict) else {}

def unique_keep_order(items):
    seen = set()
    out = []
    for x in items or []:
        x = str(x).strip()
        if not x:
            continue
        if x not in seen:
            out.append(x)
            seen.add(x)
    return out

def clamp_list(items, n):
    items = [str(x).strip() for x in (items or []) if str(x).strip()]
    if len(items) <= n:
        return items, 0
    return items[:n], len(items) - n

def one_line(s: str, width: int) -> str:
    wrapped = textwrap.wrap(str(s), width=width)
    return wrapped[0] if wrapped else str(s)

def _mpl_font_setup():
    """
    让 Matplotlib 在 Cloud 也尽量显示中文：
    - 尝试加载仓库里的 NotoSansSC-Regular.ttf
    """
    import matplotlib as mpl
    from pathlib import Path
    from matplotlib import font_manager as fm

    root = Path(__file__).resolve().parents[1]
    candidates = [
        root / "NotoSansSC-Regular.ttf",
        root / "assets" / "NotoSansSC-Regular.ttf",
        root / "fonts" / "NotoSansSC-Regular.ttf",
        root / "assets" / "fonts" / "NotoSansSC-Regular.ttf",
    ]

    font_name = None
    for p in candidates:
        if p.exists():
            try:
                fm.fontManager.addfont(str(p))
                prop = fm.FontProperties(fname=str(p))
                font_name = prop.get_name()
                break
            except Exception:
                pass

    if font_name:
        mpl.rcParams["font.sans-serif"] = [
            font_name, "Noto Sans CJK SC", "Source Han Sans SC",
            "Microsoft YaHei", "SimHei", "Arial Unicode MS", "DejaVu Sans"
        ]
    else:
        mpl.rcParams["font.sans-serif"] = [
            "Noto Sans CJK SC", "Source Han Sans SC",
            "Microsoft YaHei", "SimHei", "Arial Unicode MS", "DejaVu Sans"
        ]

    mpl.rcParams["axes.unicode_minus"] = False

def safe_radius(items, base=2.25, scale=0.03):
    n = len([x for x in (items or []) if str(x).strip()])
    return max(base, base + n * scale)

def _pick_intersection_list(intersections: dict, keys):
    for k in keys:
        v = intersections.get(k)
        if isinstance(v, list) and v:
            return v
    v0 = intersections.get(keys[0], [])
    return v0 if isinstance(v0, list) else []

# =======================
# 标题自动换行（英文两行）
# =======================
def draw_auto_title(ax, main_title, subtitle, signature, y_top, is_english, mode):
    if mode == "share":
        main_fs, sub_fs, sig_fs = 28, 18, 12
    else:
        main_fs, sub_fs, sig_fs = 24, 16, 12

    lines = []
    if is_english:
        words = main_title.split(" ")
        cur = ""
        for w in words:
            if len(cur) + len(w) + (1 if cur else 0) <= 18:
                cur = f"{cur} {w}".strip()
            else:
                if cur:
                    lines.append(cur)
                cur = w
        if cur:
            lines.append(cur)
        if len(lines) > 2:
            lines = [" ".join(lines[:-1]), lines[-1]]
    else:
        lines = [main_title]

    y = y_top - 0.75
    for line in lines:
        ax.text(0, y, line, ha="center", va="center", fontsize=main_fs, fontweight="bold")
        y -= 0.85

    ax.text(0, y - 0.10, subtitle, ha="center", va="center", fontsize=sub_fs, fontweight="bold")
    ax.text(0, y - 0.80, signature, ha="center", va="center", fontsize=sig_fs, color="#555", alpha=0.60)

# =======================
# Life Circle 海报渲染（含两两交集 + 三清单）
# =======================
def render_life_circle_png(
    canvas: str,   # preview/ig_square/ig_story/xhs_3x4/xhs_4x5
    mode: str,     # share/full
    name: str,
    dream_items, resp_items, talent_items,
    intersections: dict,
    show_n_full=10,
    center_n_share=6,
):
    _mpl_font_setup()

    blue = "#4DA3FF"
    purple = "#7E57FF"
    green = "#42C77A"

    if canvas == "preview":
        dpi = 170
        figsize = (10.5, 7.5)
        fixed = False
    else:
        dpi = 220
        fixed = True
        if canvas == "ig_square":
            px = (1080, 1080)
        elif canvas == "ig_story":
            px = (1080, 1920)
        elif canvas == "xhs_3x4":
            px = (1080, 1440)
        elif canvas == "xhs_4x5":
            px = (1080, 1350)
        else:
            px = (1080, 1080)
        figsize = (px[0] / dpi, px[1] / dpi)

    fig = plt.figure(figsize=figsize, dpi=dpi)
    ax = plt.gca()
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_position([0, 0, 1, 1])

    x_min, x_max = -6.6, 6.6
    y_min, y_max = -5.3, 7.6
    ax.set_xlim(x_min, x_max)
    ax.set_ylim(y_min, y_max)

    Dream_xy = (-1.85, -1.15)
    Talent_xy = (1.85, -1.15)
    Resp_xy = (0.0, 1.65)

    r_dream = max(2.35, safe_radius(dream_items))
    r_talent = max(2.35, safe_radius(talent_items))
    r_resp = max(2.35, safe_radius(resp_items))

    alpha_circle = 0.22 if mode == "share" else 0.26

    # 圈：紫→蓝/绿（让左右圈更显眼）
    ax.add_patch(Circle(Resp_xy,  r_resp,  color=purple, alpha=alpha_circle, lw=2, zorder=1))
    ax.add_patch(Circle(Dream_xy, r_dream, color=blue,   alpha=alpha_circle, lw=2, zorder=2))
    ax.add_patch(Circle(Talent_xy,r_talent,color=green,  alpha=alpha_circle, lw=2, zorder=2))

    is_en = st.session_state.get("lang", "zh") == "en"
    if is_en:
        title_main = "Find Your 2026 Breakthrough"
        dream_label, talent_label, resp_label = "Dream", "Talent", "Responsibility"
        center_title = "Breakthrough (Center)"
    else:
        title_main = "找到2026年人生突破点"
        dream_label, talent_label, resp_label = "梦想", "天赋", "责任"
        center_title = "三者交汇（突破点）"

    signature = f"{(name or 'YourName')} · 2026 · Life Circle"
    draw_auto_title(ax, title_main, "Life Circle", signature, y_top=y_max, is_english=is_en, mode=mode)

    # 标签：梦想/天赋底部，责任右侧（英文竖排）
    label_fs = 18
    bottom_label_y = Dream_xy[1] - r_dream - 0.55
    ax.text(Dream_xy[0], bottom_label_y, dream_label, ha="center", va="center", fontsize=label_fs, fontweight="bold")
    ax.text(Talent_xy[0], bottom_label_y, talent_label, ha="center", va="center", fontsize=label_fs, fontweight="bold")

    resp_y = Resp_xy[1] + 0.10
    ideal_x = Resp_xy[0] + r_resp + 0.55
    if is_en:
        resp_x = min(ideal_x, x_max - 0.35)
        ax.text(resp_x, resp_y, resp_label, ha="center", va="center", fontsize=label_fs, fontweight="bold", rotation=90)
    else:
        resp_x = min(ideal_x, x_max - 1.2)
        ax.text(resp_x, resp_y, resp_label, ha="left", va="center", fontsize=label_fs, fontweight="bold")

    # slogan
    ax.text(0, y_min + 0.20, "Mission → Action → Reality",
            ha="center", va="center", fontsize=13, color="#666", alpha=0.55)

    # center
    center = intersections.get("center", []) or intersections.get("中心", []) or []
    show_center, more_center = clamp_list(center, center_n_share if mode == "share" else min(10, show_n_full))
    center_lines = [f"• {one_line(x, 18)}" for x in show_center]
    if more_center > 0:
        center_lines.append(f"… {more_center} more" if is_en else f"… 还有 {more_center} 条")

    center_text = center_title + "\n" + (
        "\n".join(center_lines) if center_lines else ("(empty)" if is_en else "（空）")
    )

    ax.text(
        0.0, 0.20,
        center_text,
        ha="center", va="center",
        fontsize=13 if mode == "share" else 12,
        fontweight="bold",
        bbox=dict(
            boxstyle="round,pad=0.55,rounding_size=0.15",
            facecolor="white",
            edgecolor="#333",
            linewidth=1.1,
            alpha=0.84,
        ),
        zorder=6
    )

    # Full：三清单 + 三交集
    if mode == "full":
        def _list_block(title, items, x, y):
            show, more = clamp_list(items, 7)
            lines = [f"• {one_line(s, 18)}" for s in show]
            if more > 0:
                lines.append(f"… {more} more" if is_en else f"… 还有 {more} 条")
            txt = title + "\n" + ("\n".join(lines) if lines else ("(empty)" if is_en else "（空）"))
            ax.text(
                x, y, txt,
                ha="center", va="center",
                fontsize=10,
                bbox=dict(boxstyle="round,pad=0.30,rounding_size=0.12",
                          facecolor="white", edgecolor="#999", linewidth=0.8, alpha=0.55),
                zorder=5
            )

        _list_block("Responsibility List" if is_en else "责任清单", resp_items, Resp_xy[0], Resp_xy[1] + 0.75)
        _list_block("Dream List" if is_en else "梦想清单", dream_items, Dream_xy[0] - 0.25, Dream_xy[1] + 0.15)
        _list_block("Talent List" if is_en else "天赋清单", talent_items, Talent_xy[0] + 0.25, Talent_xy[1] + 0.15)

        resp_dream = _pick_intersection_list(intersections, ["resp_dream", "责任∩梦想", "rd"])
        resp_talent = _pick_intersection_list(intersections, ["resp_talent", "责任∩天赋", "rt"])
        dream_talent = _pick_intersection_list(intersections, ["dream_talent", "梦想∩天赋", "dt"])

        def _fmt_block(title, items, max_n=4):
            show, more = clamp_list(items, max_n)
            lines = [f"• {one_line(x, 14)}" for x in show]
            if more > 0:
                lines.append(f"… {more} more" if is_en else f"… 还有 {more} 条")
            return title + "\n" + ("\n".join(lines) if lines else ("(empty)" if is_en else "（空）"))

        ax.text(-3.10, 0.95, _fmt_block("Resp ∩ Dream" if is_en else "责任 ∩ 梦想", resp_dream),
                ha="center", va="center", fontsize=9,
                bbox=dict(boxstyle="round,pad=0.25", facecolor="white", edgecolor="#AAA", alpha=0.60),
                zorder=7)
        ax.text(3.10, 0.95, _fmt_block("Resp ∩ Talent" if is_en else "责任 ∩ 天赋", resp_talent),
                ha="center", va="center", fontsize=9,
                bbox=dict(boxstyle="round,pad=0.25", facecolor="white", edgecolor="#AAA", alpha=0.60),
                zorder=7)
        ax.text(0.0, -2.75, _fmt_block("Dream ∩ Talent" if is_en else "梦想 ∩ 天赋", dream_talent),
                ha="center", va="center", fontsize=9,
                bbox=dict(boxstyle="round,pad=0.25", facecolor="white", edgecolor="#AAA", alpha=0.60),
                zorder=7)

    buf = io.BytesIO()
    if fixed:
        fig.savefig(buf, format="png", dpi=dpi, facecolor="white")
    else:
        fig.savefig(buf, format="png", dpi=dpi, facecolor="white", bbox_inches="tight")
    b = buf.getvalue()
    buf.close()
    plt.close(fig)
    return b

# =======================
# 36×10 Excel 导出：6×6 大表（基于 store.py 的 sprints/tasks）
# =======================
def build_36x10_excel() -> bytes:
    periods = get_sprints()
    if not periods or len(periods) < 1:
        return b""

    # store.py 的 sprint 是 dict，字段 start_date/end_date 是 isoformat 字符串
    # 这里不依赖日期对象，直接容错
    safe_periods = []
    for p in periods:
        if not isinstance(p, dict):
            continue
        if not p.get("sprint_no"):
            continue
        safe_periods.append(p)

    if not safe_periods:
        return b""

    BLOCK_COLS = 3
    BLOCK_ROWS = 10
    GAP_COL = 1
    GAP_ROW = 1

    wb = Workbook()
    ws = wb.active
    ws.title = "36×10 Plan" if st.session_state.get("lang", "zh") == "en" else "36×10 自我提升计划"

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    font_header = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    font_body = Font(name="Microsoft YaHei", size=10, color="1F1F1F")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    fill_header = PatternFill("solid", fgColor="6C5CE7")
    fill_obj = PatternFill("solid", fgColor="F7F7FB")
    fill_task = PatternFill("solid", fgColor="FFFFFF")
    fill_done = PatternFill("solid", fgColor="E9F7EF")

    total_cols = 6 * (BLOCK_COLS + GAP_COL)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_text = "36×10 Growth Plan (6×6 Master Sheet)" if st.session_state.get("lang", "zh") == "en" else "36×10 自我提升计划（6×6 大表）"
    tcell = ws.cell(row=1, column=1, value=title_text)
    tcell.font = Font(name="Microsoft YaHei", bold=True, size=16, color="1F1F1F")
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for c in range(1, total_cols + 1):
        letter = get_column_letter(c)
        if (c % (BLOCK_COLS + GAP_COL)) == 0:
            ws.column_dimensions[letter].width = 3
        else:
            ws.column_dimensions[letter].width = 18

    for r in range(2, 2 + 6 * (BLOCK_ROWS + GAP_ROW) + 2):
        ws.row_dimensions[r].height = 18

    periods_sorted = sorted(safe_periods, key=lambda x: int(x.get("sprint_no", 0)))

    def top_left_of_block(sprint_no: int):
        idx = sprint_no - 1
        block_r = idx // 6
        block_c = idx % 6
        start_row = 2 + block_r * (BLOCK_ROWS + GAP_ROW)
        start_col = 1 + block_c * (BLOCK_COLS + GAP_COL)
        return start_row, start_col

    def merge_block(row, col, r_span, c_span):
        ws.merge_cells(start_row=row, start_column=col, end_row=row + r_span - 1, end_column=col + c_span - 1)

    def set_block_border(r0, c0, r1, c1):
        for rr in range(r0, r1 + 1):
            for cc in range(c0, c1 + 1):
                ws.cell(rr, cc).border = border

    deliverable_label = "Deliverables:\n" if st.session_state.get("lang", "zh") == "en" else "交付物/成果：\n"

    for sp in periods_sorted:
        sprint_no = int(sp.get("sprint_no", 0))
        if sprint_no <= 0:
            continue

        r0, c0 = top_left_of_block(sprint_no)
        r1 = r0 + BLOCK_ROWS - 1
        c1 = c0 + BLOCK_COLS - 1

        theme = (sp.get("theme") or "").strip()
        header_text = theme if theme else ("Untitled" if st.session_state.get("lang", "zh") == "en" else "未命名主题")
        header_text = (f"Cycle {sprint_no} | {header_text}" if st.session_state.get("lang", "zh") == "en"
                       else f"第{sprint_no}周期｜{header_text}")

        merge_block(r0, c0, 1, BLOCK_COLS)
        hc = ws.cell(r0, c0, header_text)
        hc.font = font_header
        hc.fill = fill_header
        hc.alignment = align_center
        ws.row_dimensions[r0].height = 26

        obj = (sp.get("objective") or "").strip()
        obj_text = obj if obj else ("(Not set)" if st.session_state.get("lang", "zh") == "en" else "（未填写交付物）")
        merge_block(r0 + 1, c0, 2, BLOCK_COLS)
        oc = ws.cell(r0 + 1, c0, f"{deliverable_label}{obj_text}")
        oc.font = font_body
        oc.fill = fill_obj
        oc.alignment = align_left
        ws.row_dimensions[r0 + 1].height = 38
        ws.row_dimensions[r0 + 2].height = 38

        tasks = list_tasks_for_sprint(sprint_no) or []
        max_tasks = 6
        show = tasks[:max_tasks]
        more = max(0, len(tasks) - len(show))

        for i in range(max_tasks):
            rr = r0 + 3 + i
            merge_block(rr, c0, 1, BLOCK_COLS)
            if i < len(show):
                tt = show[i]
                done = bool(tt.get("done")) if isinstance(tt, dict) else False
                title = (tt.get("title") if isinstance(tt, dict) else "") or ""
                mark = "✅" if done else "⬜"
                txt = f"{mark} {title}"
                cell = ws.cell(rr, c0, txt)
                cell.fill = fill_done if done else fill_task
            else:
                cell = ws.cell(rr, c0, "")
                cell.fill = fill_task

            cell.font = font_body
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[rr].height = 20

        rr_hint = r0 + 9
        merge_block(rr_hint, c0, 1, BLOCK_COLS)
        hint = (f"… {more} more tasks" if st.session_state.get("lang", "zh") == "en" else f"…还有 {more} 条任务") if more > 0 else ""
        hint_cell = ws.cell(rr_hint, c0, hint)
        hint_cell.font = Font(name="Microsoft YaHei", size=9, color="666666", italic=True)
        hint_cell.alignment = Alignment(horizontal="right", vertical="center")
        hint_cell.fill = fill_task
        ws.row_dimensions[rr_hint].height = 18

        set_block_border(r0, c0, r1, c1)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# =======================
# 读取数据（Annual Dig）
# =======================
dig = get_or_create_annual_dig()
talent = safe_load_json(getattr(dig, "talent_json", "{}"))
resp = safe_load_json(getattr(dig, "responsibility_json", "{}"))
dream = safe_load_json(getattr(dig, "dream_json", "{}"))
inter = safe_load_json(getattr(dig, "intersections_json", "{}"))

name = (get_meta(inter).get("name", "") or "").strip()

def _sum_quadrants(d):
    if not isinstance(d, dict):
        return []
    all_items = []
    for v in d.values():
        if isinstance(v, list):
            all_items.extend(v)
    return unique_keep_order(all_items)

dream_items = _sum_quadrants(dream)
resp_items = _sum_quadrants(resp)
talent_items = _sum_quadrants(talent)

# ============================================================
# ✅ 1) JSON 备份（默认折叠，不影响产品感）
# ============================================================
with st.expander("📦 数据备份 / Backup（建议先下载保存）", expanded=False):
    st.caption("用于内测阶段：避免刷新/换设备后数据丢失。下载的 JSON 可随时上传恢复。")
    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            "⬇️ 下载我的数据（JSON）",
            data=export_user_json(),
            file_name="bright_future_backup.json",
            mime="application/json",
            use_container_width=True,
        )
    with c2:
        up = st.file_uploader("⬆️ 上传继续编辑（JSON）", type=["json"])
        if up is not None:
            import_user_json(up.getvalue())
            st.success("已导入 ✅")
            st.rerun()

# ============================================================
# A | 海报导出
# ============================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.subheader(t("poster_section"))
st.caption("分享版适合发布；完整版适合存档/复盘。" if st.session_state.get("lang", "zh") == "zh"
           else "Share mode is great for posting; Full mode is better for archive/review.")

mode_ui = st.radio(
    t("mode_label"),
    [t("mode_share"), t("mode_full")],
    horizontal=True,
    index=0,
)
mode_key = "share" if mode_ui == t("mode_share") else "full"

preview_png = render_life_circle_png(
    canvas="preview",
    mode=mode_key,
    name=name,
    dream_items=dream_items,
    resp_items=resp_items,
    talent_items=talent_items,
    intersections=inter,
)

st.image(preview_png, width=1100)

suffix = "share" if mode_key == "share" else "full"
base_name = f"{(name or 'YourName')}_2026_LifeCircle_{suffix}"

c1, c2, c3, c4 = st.columns(4)

def export_poster(canvas_key: str) -> bytes:
    return render_life_circle_png(
        canvas=canvas_key,
        mode=mode_key,
        name=name,
        dream_items=dream_items,
        resp_items=resp_items,
        talent_items=talent_items,
        intersections=inter,
    )

with c1:
    st.download_button(t("download_ig_square"), export_poster("ig_square"),
                       file_name=f"{base_name}_IG_1x1.png", mime="image/png", use_container_width=True)
with c2:
    st.download_button(t("download_ig_story"), export_poster("ig_story"),
                       file_name=f"{base_name}_IG_9x16.png", mime="image/png", use_container_width=True)
with c3:
    st.download_button(t("download_xhs_3x4"), export_poster("xhs_3x4"),
                       file_name=f"{base_name}_3x4.png", mime="image/png", use_container_width=True)
with c4:
    st.download_button(t("download_xhs_4x5"), export_poster("xhs_4x5"),
                       file_name=f"{base_name}_4x5.png", mime="image/png", use_container_width=True)

st.markdown("</div>", unsafe_allow_html=True)

# ============================================================
# B | Excel 导出
# ============================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.subheader(t("excel_section"))
st.caption("每格一个10天行动周期：表头=主题，下面=交付物，再下面=任务列表（含完成状态）。"
           if st.session_state.get("lang", "zh") == "zh"
           else "Each block is a 10-day cycle: header=theme, then deliverables, then tasks (with done status).")

xlsx_bytes = build_36x10_excel()

if not xlsx_bytes:
    st.info("还没有生成 36×10 行动周期。请先到「② 36×10」页面生成周期，再回来导出。"
            if st.session_state.get("lang", "zh") == "zh"
            else "No 36×10 cycles yet. Please generate them on page ② first.")
else:
    xlsx_name = (
        f"{(name or 'YourName')}_36x10_plan.xlsx"
        if st.session_state.get("lang", "zh") == "en"
        else f"{(name or 'YourName')}_36x10_自我提升计划.xlsx"
    )
    st.download_button(
        t("download_excel"),
        data=xlsx_bytes,
        file_name=xlsx_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

st.markdown("</div>", unsafe_allow_html=True)
